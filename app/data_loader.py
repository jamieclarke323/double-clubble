"""
Loads the "Double Clubble" quiz data straight from the source spreadsheet.

Spreadsheet layout (sheet "DoubleClubble"):
    The sheet is made up of side-by-side blocks, one per Charlton Athletic
    away fixture in the 2026/27 Championship season. Each block is 9 columns
    wide:

        col 0: Player name
        col 1: Position
        col 2: Years played for Charlton
        col 3: Appearances for Charlton
        col 4: Goals for Charlton
        col 5: Years played for the opposition club
        col 6: Appearances for the opposition club
        col 7: Goals for the opposition club
        col 8: (spacer column, always blank)

    Row 1 of each block holds a group label ("CharltonWestHam") in column 0
    and the fixture date (e.g. "22-08-2026") in column 4.
    Row 2 holds "Charlton"/"Charlton Athletic" in column 2 and the full
    opposition club name in column 5.
    Row 3 holds the column headers ("Player", "Position", "Years", ...).
    Player rows start on row 4 and run until the block runs out of players
    (short blocks are simply left blank - we stop at the first blank name).

There is also an "AwayFixtures" sheet with a clean Location / Home Team /
Date / Away Team / Result table, which we use to enrich the venue for each
fixture on the home page.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl

# Charlton play their home 7am unlock rule in UK local time.
QUIZ_TIMEZONE = ZoneInfo("Europe/London")
UNLOCK_HOUR = 7

SHEET_NAME = "DoubleClubble"
FIXTURES_SHEET_NAME = "AwayFixtures"
BLOCK_WIDTH = 9
FIRST_PLAYER_ROW = 4
MAX_PLAYER_ROW = 200  # generous upper bound; real data stops far earlier

# A handful of players in the source data have surnames made up of more than
# one word (e.g. "Tal Ben Haim" -> "Ben Haim"). Simple "last word" splitting
# would produce the wrong (or an overly obscure) answer for these, so they
# are called out explicitly. The single trailing word is still accepted as
# an alternative answer (see Player.acceptable_answers).
MULTI_WORD_SURNAMES = {
    "Tal Ben Haim": "Ben Haim",
    "Paolo Di Canio": "Di Canio", 
    "Ricardo Vaz Tê": "Vaz Tê",
    "Tahar El Khalej": "El Khalej",
    "Jesurun Rak Sakyi": "Rak Sakyi",
    "Jesurun Rak-Sakyi": "Rak Sakyi",
}


def _derive_surname(full_name: str) -> str:
    """Best-effort surname extraction for a player's full name."""
    full_name = full_name.strip()
    if full_name in MULTI_WORD_SURNAMES:
        return MULTI_WORD_SURNAMES[full_name]
    parts = full_name.split()
    return parts[-1] if parts else full_name


def _acceptable_answers(full_name: str, surname: str) -> list[str]:
    """All answer strings that should be treated as fully correct."""
    answers = {surname, full_name}
    # Also accept the plain final word even when we picked a multi-word
    # surname above (e.g. "Haim" as well as "Ben Haim").
    parts = full_name.split()
    if parts:
        answers.add(parts[-1])
    return sorted(answers, key=len, reverse=True)


def _derive_initials(full_name: str) -> str:
    """Initials clue, e.g. 'Tal Ben Haim' -> 'T.B.H.'."""
    parts = full_name.split()
    return "".join(f"{part[0].upper()}." for part in parts if part)


@dataclass
class Player:
    name: str
    position: str
    years_charlton: str
    apps_charlton: str | int
    goals_charlton: str | int
    years_opponent: str
    apps_opponent: str | int
    goals_opponent: str | int
    surname: str = field(init=False)
    acceptable_answers: list[str] = field(init=False)
    initials: str = field(init=False)

    def __post_init__(self) -> None:
        self.surname = _derive_surname(self.name)
        self.acceptable_answers = _acceptable_answers(self.name, self.surname)
        self.initials = _derive_initials(self.name)

    def clue_value(self, clue_type: str) -> str:
        if clue_type == "apps":
            return f"Charlton: {self.apps_charlton} / {self.opponent_apps_label}: {self.apps_opponent}"
        if clue_type == "years":
            return f"Charlton: {self.years_charlton} / {self.opponent_apps_label}: {self.years_opponent}"
        if clue_type == "goals":
            return f"Charlton: {self.goals_charlton} / {self.opponent_apps_label}: {self.goals_opponent}"
        if clue_type == "position":
            return self.position
        if clue_type == "initials":
            return self.initials
        raise ValueError(f"Unknown clue type: {clue_type}")

    # populated by Quiz.__post_init__ so clues can reference the opponent name
    opponent_apps_label: str = field(default="Opponent", init=False)


@dataclass
class Quiz:
    quiz_id: str
    opponent: str
    fixture_date: date
    players: list[Player]
    venue: str | None = None
    home_team: str | None = None
    unlock_at: datetime = field(init=False)

    def __post_init__(self) -> None:
        self.unlock_at = datetime.combine(
            self.fixture_date, time(hour=UNLOCK_HOUR), tzinfo=QUIZ_TIMEZONE
        )
        for player in self.players:
            player.opponent_apps_label = self.opponent

    def is_unlocked(self, now: datetime | None = None) -> bool:
        now = now or datetime.now(QUIZ_TIMEZONE)
        if now.tzinfo is None:
            now = now.replace(tzinfo=QUIZ_TIMEZONE)
        return now >= self.unlock_at

    @property
    def total_players(self) -> int:
        return len(self.players)

    @property
    def unlock_display(self) -> str:
        """Human readable unlock time, safe across platforms (no %-I)."""
        hour_12 = self.unlock_at.strftime("%I").lstrip("0") or "12"
        return (
            f"{self.unlock_at.strftime('%A %d %B %Y')} at "
            f"{hour_12}:{self.unlock_at.strftime('%M%p')}"
        )


def _slugify(name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    return slug


def _parse_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value).strip(), "%d-%m-%Y").date()


def _cell(ws, row: int, col: int):
    return ws.cell(row=row, column=col).value


def _clean_number(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _clean_years(value) -> str:
    """Format a Years cell as a string, e.g. 2014.0 -> "2014", but keep
    year-range strings such as "1997-1998" untouched."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _load_fixture_extras(ws) -> dict[str, dict[str, str]]:
    """Map "opponent name" -> {venue, home_team} from the AwayFixtures sheet."""
    extras: dict[str, dict[str, str]] = {}
    for row in range(2, ws.max_row + 1):
        venue = _cell(ws, row, 1)
        home_team = _cell(ws, row, 2)
        fixture_date = _cell(ws, row, 3)
        if not home_team or not fixture_date:
            continue
        extras[str(fixture_date).strip()] = {
            "venue": venue,
            "home_team": home_team,
        }
    return extras


def load_quizzes(xlsx_path: str | Path) -> list[Quiz]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME]
    fixtures_ws = wb[FIXTURES_SHEET_NAME] if FIXTURES_SHEET_NAME in wb.sheetnames else None
    extras_by_date = _load_fixture_extras(fixtures_ws) if fixtures_ws is not None else {}

    quizzes: list[Quiz] = []
    for start_col in range(1, ws.max_column + 1, BLOCK_WIDTH):
        date_raw = _cell(ws, 1, start_col + 4)
        opponent = _cell(ws, 2, start_col + 5)
        if not date_raw or not opponent:
            break
        opponent = str(opponent).strip()
        fixture_date = _parse_date(date_raw)

        players: list[Player] = []
        for row in range(FIRST_PLAYER_ROW, MAX_PLAYER_ROW):
            name = _cell(ws, row, start_col)
            if not name:
                break
            players.append(
                Player(
                    name=str(name).strip(),
                    position=str(_cell(ws, row, start_col + 1) or ""),
                    years_charlton=_clean_years(_cell(ws, row, start_col + 2)),
                    apps_charlton=_clean_number(_cell(ws, row, start_col + 3)),
                    goals_charlton=_clean_number(_cell(ws, row, start_col + 4)),
                    years_opponent=_clean_years(_cell(ws, row, start_col + 5)),
                    apps_opponent=_clean_number(_cell(ws, row, start_col + 6)),
                    goals_opponent=_clean_number(_cell(ws, row, start_col + 7)),
                )
            )

        extras = extras_by_date.get(str(date_raw).strip(), {})
        quizzes.append(
            Quiz(
                quiz_id=_slugify(opponent),
                opponent=opponent,
                fixture_date=fixture_date,
                players=players,
                venue=extras.get("venue"),
                home_team=extras.get("home_team"),
            )
        )

    quizzes.sort(key=lambda q: q.fixture_date)
    return quizzes
